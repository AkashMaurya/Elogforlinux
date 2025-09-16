from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db import transaction, models
from django.views.decorators.http import require_POST
import os
import csv
import io
import json
from datetime import timedelta
from django.contrib.auth.hashers import make_password

# Models
from admin_section.models import *
from accounts.models import CustomUser, Student, Doctor, Staff
from student_section.models import SupportTicket, StudentLogFormModel
from doctor_section.models import DoctorSupportTicket

# Forms
from .forms import BulkUserUploadForm, CSVUploadForm, BlogForm, BlogCategoryForm
from student_section.forms import AdminResponseForm
from doctor_section.forms import AdminDoctorResponseForm, BatchReviewForm, LogReviewForm
from django.core.paginator import Paginator

# Django predefined models
from django.db.models import Count, Q
from django.db.models.functions import TruncMonth
from datetime import datetime
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
import tablib
from utils.pdf_utils import add_agu_header, get_common_styles, add_footer_info

@login_required
def department_report(request):
    """View for Department Report"""
    if request.user.role != 'admin':
        messages.error(request, "You don't have permission to access this page.")
        return redirect('admin_section:admin_dash')

    # Get all departments with related data
    departments = Department.objects.all().order_by('name')

    # Get filter parameters
    department_filter = request.GET.get('department')
    year_filter = request.GET.get('year')
    section_filter = request.GET.get('section')

    # Base queryset for logs
    logs = StudentLogFormModel.objects.select_related('student', 'department', 'activity_type', 'training_site')

    # Apply filters if provided
    if department_filter:
        logs = logs.filter(department_id=department_filter)
        # Get the specific department for filtering
        selected_department = departments.filter(id=department_filter).first()
        if selected_department:
            departments = departments.filter(id=department_filter)
    if year_filter:
        logs = logs.filter(log_year_id=year_filter)
    if section_filter:
        logs = logs.filter(log_year_section_id=section_filter)

    # Calculate department statistics
    department_stats = []
    for dept in departments:
        dept_logs = logs.filter(department=dept)
        total_logs = dept_logs.count()
        reviewed_logs = dept_logs.filter(is_reviewed=True).count()
        pending_logs = total_logs - reviewed_logs
        approved_logs = dept_logs.filter(is_reviewed=True).exclude(reviewer_comments__startswith='REJECTED').count()
        rejected_logs = dept_logs.filter(is_reviewed=True, reviewer_comments__startswith='REJECTED').count()

        department_stats.append({
            'department': dept,
            'total_logs': total_logs,
            'reviewed_logs': reviewed_logs,
            'pending_logs': pending_logs,
            'approved_logs': approved_logs,
            'rejected_logs': rejected_logs,
            'doctors_count': dept.doctors.count(),
            'students_count': Student.objects.filter(group__log_year_section=dept.log_year_section).count() if dept.log_year_section else 0
        })

    # Prepare chart data
    import json
    from django.db.models import Count

    # Case Types Data (Activity Types) - Real data from logs
    case_types = logs.values('activity_type__name').annotate(count=Count('id')).order_by('-count')
    case_types_data = {
        'labels': [item['activity_type__name'] or 'Unknown' for item in case_types],
        'data': [item['count'] for item in case_types]
    }

    # Training Sites Data - Real data from logs
    training_sites = logs.values('training_site__name').annotate(count=Count('id')).order_by('-count')
    training_sites_data = {
        'labels': [item['training_site__name'] or 'Unknown' for item in training_sites],
        'data': [item['count'] for item in training_sites]
    }

    # Activity Types Data (same as case types but can be filtered)
    activity_types_data = case_types_data.copy()

    # Participation Data - Real data from logs
    participation_logs = logs.values('participation_type').annotate(count=Count('id')).order_by('-count')
    participation_data = {
        'labels': [item['participation_type'] or 'Not Specified' for item in participation_logs],
        'data': [item['count'] for item in participation_logs]
    }

    # Monthly Data - Real data from logs
    monthly_logs = logs.annotate(month=TruncMonth('date')).values('month').annotate(count=Count('id')).order_by('month')
    monthly_data = {
        'labels': [item['month'].strftime('%B %Y') if item['month'] else 'Unknown' for item in monthly_logs],
        'data': [item['count'] for item in monthly_logs]
    }

    # Core Diagnosis Data - Additional chart data
    core_diagnosis_logs = logs.values('core_diagnosis__name').annotate(count=Count('id')).order_by('-count')[:10]  # Top 10
    core_diagnosis_data = {
        'labels': [item['core_diagnosis__name'] or 'Unknown' for item in core_diagnosis_logs],
        'data': [item['count'] for item in core_diagnosis_logs]
    }

    # Patient Gender Data - Additional chart data
    gender_logs = logs.values('patient_gender').annotate(count=Count('id')).order_by('-count')
    gender_data = {
        'labels': [item['patient_gender'] or 'Not Specified' for item in gender_logs],
        'data': [item['count'] for item in gender_logs]
    }

    # Department-wise case distribution
    dept_case_logs = logs.values('department__name').annotate(count=Count('id')).order_by('-count')
    dept_case_data = {
        'labels': [item['department__name'] or 'Unknown' for item in dept_case_logs],
        'data': [item['count'] for item in dept_case_logs]
    }

    # Approval Status Data
    total_logs_count = logs.count()
    reviewed_count = logs.filter(is_reviewed=True).count()
    pending_count = total_logs_count - reviewed_count
    approved_count = logs.filter(is_reviewed=True).exclude(reviewer_comments__startswith='REJECTED').count()
    rejected_count = logs.filter(is_reviewed=True, reviewer_comments__startswith='REJECTED').count()

    approval_status_data = {
        'approved': approved_count,
        'pending': pending_count,
        'rejected': rejected_count
    }

    # Get years and sections for filters
    years = LogYear.objects.all().order_by('-year_name')
    sections = LogYearSection.objects.all().order_by('year_section_name')
    all_departments = Department.objects.all().order_by('name')

    # Calculate totals
    total_doctors = Doctor.objects.filter(departments__in=departments).distinct().count() if department_filter else Doctor.objects.count()
    total_training_sites = TrainingSite.objects.count()
    total_activity_types = ActivityType.objects.filter(department__in=departments).count() if department_filter else ActivityType.objects.count()

    context = {
        'department_stats': department_stats,
        'departments': all_departments,
        'years': years,
        'sections': sections,
        'selected_department': department_filter,
        'selected_year': year_filter,
        'selected_section': section_filter,
        'total_departments': departments.count(),
        'total_logs': logs.count(),
        'total_doctors': total_doctors,
        'total_training_sites': total_training_sites,
        'total_activity_types': total_activity_types,
        # Chart data as JSON
        'case_types_data': json.dumps(case_types_data),
        'training_sites_data': json.dumps(training_sites_data),
        'activity_types_data': json.dumps(activity_types_data),
        'participation_data': json.dumps(participation_data),
        'monthly_data': json.dumps(monthly_data),
        'approval_status_data': json.dumps(approval_status_data),
        'core_diagnosis_data': json.dumps(core_diagnosis_data),
        'gender_data': json.dumps(gender_data),
        'dept_case_data': json.dumps(dept_case_data),
    }

    return render(request, 'admin_section/department_report.html', context)


@login_required
def department_report_export(request):
    """Export Department Report as PDF or Excel"""
    if request.user.role != 'admin':
        messages.error(request, "You don't have permission to access this page.")
        return redirect('admin_section:admin_dash')

    export_format = request.GET.get('format', 'pdf')

    # Get the same data as the main report
    departments = Department.objects.all()

    # Get filter parameters
    selected_department = request.GET.get('department')
    selected_year = request.GET.get('year')

    # Filter departments if specific department selected
    if selected_department:
        departments = departments.filter(id=selected_department)

    # Prepare data for export
    department_data = []
    for dept in departments:
        # Get logs for this department (students are connected to departments through logs)
        logs = StudentLogFormModel.objects.filter(department=dept)
        if selected_year:
            logs = logs.filter(student__group__log_year__year_name=selected_year)

        # Get unique students who have logs in this department
        student_ids = logs.values_list('student_id', flat=True).distinct()
        students = Student.objects.filter(id__in=student_ids)
        if selected_year:
            students = students.filter(group__log_year__year_name=selected_year)

        # Calculate statistics
        total_students = students.count()
        total_logs = logs.count()
        reviewed_logs = logs.filter(is_reviewed=True).count()
        pending_logs = total_logs - reviewed_logs

        department_data.append({
            'name': dept.name,
            'total_students': total_students,
            'total_logs': total_logs,
            'reviewed_logs': reviewed_logs,
            'pending_logs': pending_logs,
            'review_rate': f"{(reviewed_logs/total_logs*100):.1f}%" if total_logs > 0 else "0%"
        })

    if export_format == 'excel':
        return export_department_excel(department_data, selected_department, selected_year)
    else:
        return export_department_pdf(department_data, selected_department, selected_year)


def export_department_excel(department_data, selected_department=None, selected_year=None):
    """Export department report to Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, Reference
    from django.http import HttpResponse
    import io

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Department Report"

    # Add title
    title = "Department Report"
    if selected_department:
        dept_name = Department.objects.get(id=selected_department).name
        title += f" - {dept_name}"
    if selected_year:
        title += f" - {selected_year}"

    ws['A1'] = title
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:F1')

    # Add headers
    headers = ['Department', 'Total Students', 'Total Logs', 'Reviewed Logs', 'Pending Logs', 'Review Rate']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Add data
    for row, dept in enumerate(department_data, 4):
        ws.cell(row=row, column=1, value=dept['name'])
        ws.cell(row=row, column=2, value=dept['total_students'])
        ws.cell(row=row, column=3, value=dept['total_logs'])
        ws.cell(row=row, column=4, value=dept['reviewed_logs'])
        ws.cell(row=row, column=5, value=dept['pending_logs'])
        ws.cell(row=row, column=6, value=dept['review_rate'])

    # Add chart
    if len(department_data) > 0:
        chart = BarChart()
        chart.title = "Department Statistics"
        chart.x_axis.title = "Departments"
        chart.y_axis.title = "Count"

        # Data for chart
        data = Reference(ws, min_col=2, min_row=3, max_col=5, max_row=len(department_data) + 3)
        categories = Reference(ws, min_col=1, min_row=4, max_row=len(department_data) + 3)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Add chart to worksheet
        ws.add_chart(chart, "H3")

    # Auto-adjust column widths
    headers = ['Department', 'Total Students', 'Total Logs', 'Reviewed Logs', 'Pending Logs', 'Review Rate']
    for col_num in range(1, len(headers) + 1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_num)
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_num)
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="department_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

    wb.save(response)
    return response


def export_department_pdf(department_data, selected_department=None, selected_year=None):
    """Export department report to PDF"""
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    from django.http import HttpResponse
    import io

    # Create response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="department_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'

    # Create PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    # Add header
    add_agu_header(elements)

    # Add title
    title = "Department Report"
    if selected_department:
        dept_name = Department.objects.get(id=selected_department).name
        title += f" - {dept_name}"
    if selected_year:
        title += f" - {selected_year}"

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1,  # Center alignment
        textColor=colors.HexColor('#1f2937')
    )

    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 20))

    # Create table data
    table_data = [['Department', 'Total Students', 'Total Logs', 'Reviewed Logs', 'Pending Logs', 'Review Rate']]
    for dept in department_data:
        table_data.append([
            dept['name'],
            str(dept['total_students']),
            str(dept['total_logs']),
            str(dept['reviewed_logs']),
            str(dept['pending_logs']),
            dept['review_rate']
        ])

    # Create table
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 30))

    # Add chart if data exists
    if len(department_data) > 0:
        # Create chart
        drawing = Drawing(400, 200)
        chart = VerticalBarChart()
        chart.x = 50
        chart.y = 50
        chart.height = 125
        chart.width = 300
        chart.data = [
            [dept['total_students'] for dept in department_data],
            [dept['total_logs'] for dept in department_data],
            [dept['reviewed_logs'] for dept in department_data]
        ]
        chart.categoryAxis.categoryNames = [dept['name'][:10] + '...' if len(dept['name']) > 10 else dept['name'] for dept in department_data]
        chart.valueAxis.valueMin = 0
        chart.bars[0].fillColor = colors.HexColor('#3b82f6')
        chart.bars[1].fillColor = colors.HexColor('#10b981')
        chart.bars[2].fillColor = colors.HexColor('#f59e0b')

        drawing.add(chart)
        elements.append(drawing)

    # Add footer
    add_footer_info(elements)

    # Build PDF
    doc.build(elements)

    # Get PDF data
    pdf_data = buffer.getvalue()
    buffer.close()

    response.write(pdf_data)
    return response

@login_required
def student_report(request):
    """View for Student Report with enhanced dashboard and filtering"""
    if request.user.role != 'admin':
        messages.error(request, "You don't have permission to access this page.")
        return redirect('admin_section:admin_dash')

    # Get filter parameters
    department_filter = request.GET.get('department')
    student_filter = request.GET.get('student')
    search_query = request.GET.get('q', '').strip()

    # Base queryset for logs
    logs = StudentLogFormModel.objects.select_related(
        'student', 'student__user', 'department', 'activity_type',
        'training_site', 'core_diagnosis', 'tutor'
    ).all()

    # Apply department filter
    if department_filter:
        logs = logs.filter(department_id=department_filter)

    # Apply student filter
    if student_filter:
        logs = logs.filter(student_id=student_filter)

    # Apply search filter
    if search_query:
        logs = logs.filter(
            Q(student__user__first_name__icontains=search_query) |
            Q(student__user__last_name__icontains=search_query) |
            Q(student__user__email__icontains=search_query) |
            Q(student__student_id__icontains=search_query)
        )

    # Calculate summary statistics
    total_logs = logs.count()
    reviewed_logs = logs.filter(is_reviewed=True).count()
    pending_logs = logs.filter(is_reviewed=False).count()
    approved_logs = logs.filter(is_reviewed=True).exclude(reviewer_comments__startswith='REJECTED').count()
    rejected_logs = logs.filter(is_reviewed=True, reviewer_comments__startswith='REJECTED').count()

    # Get unique doctors - fix duplicate issue
    unique_doctor_ids = logs.values_list('tutor', flat=True).distinct()
    doctors = Doctor.objects.filter(id__in=unique_doctor_ids).select_related('user')
    doctor_names = []
    for doctor in doctors:
        full_name = f"{doctor.user.first_name} {doctor.user.last_name}".strip()
        if full_name:
            doctor_names.append(full_name)
        else:
            doctor_names.append(doctor.user.username)
    doctor_names = sorted(set(doctor_names))  # Remove any remaining duplicates and sort

    # Case Types Data (using core_diagnosis as case type)
    case_types_data = logs.values('core_diagnosis__name').annotate(
        count=Count('id')
    ).order_by('-count')

    case_types_chart = {
        'labels': [item['core_diagnosis__name'] or 'Unknown' for item in case_types_data],
        'data': [item['count'] for item in case_types_data]
    }

    # Training Sites Data
    training_sites_data = logs.values('training_site__name').annotate(
        count=Count('id')
    ).order_by('-count')

    training_sites_chart = {
        'labels': [item['training_site__name'] for item in training_sites_data],
        'data': [item['count'] for item in training_sites_data]
    }

    # Activity Types Data
    activity_types_data = logs.values('activity_type__name').annotate(
        count=Count('id')
    ).order_by('-count')

    activity_types_chart = {
        'labels': [item['activity_type__name'] for item in activity_types_data],
        'data': [item['count'] for item in activity_types_data]
    }

    # Participation Types Data
    participation_data = logs.values('participation_type').annotate(
        count=Count('id')
    ).order_by('-count')

    participation_chart = {
        'labels': [item['participation_type'] for item in participation_data],
        'data': [item['count'] for item in participation_data]
    }

    # Monthly Cases Data
    from django.db.models.functions import TruncMonth
    monthly_data = logs.annotate(
        month=TruncMonth('date')
    ).values('month').annotate(
        count=Count('id')
    ).order_by('month')

    monthly_chart = {
        'labels': [item['month'].strftime('%b %Y') for item in monthly_data],
        'data': [item['count'] for item in monthly_data]
    }

    # Approval Status Data
    approval_status_chart = {
        'approved': approved_logs,
        'pending': pending_logs,
        'rejected': rejected_logs
    }

    # Get filter options
    departments = Department.objects.all().order_by('name')
    students = Student.objects.select_related('user', 'group').all().order_by('user__first_name', 'user__last_name')

    # Initialize variables
    selected_student_obj = None
    search_results_info = None

    # Get selected student details if a specific student is selected
    if student_filter:
        try:
            selected_student_obj = Student.objects.select_related('user', 'group').get(id=student_filter)
        except Student.DoesNotExist:
            pass
    elif search_query:
        # If searching, try to find matching students
        search_students = Student.objects.select_related('user', 'group').filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(user__email__icontains=search_query) |
            Q(student_id__icontains=search_query)
        )

        # Check for exact ID match first
        exact_id_match = Student.objects.select_related('user', 'group').filter(
            student_id__iexact=search_query
        ).first()

        if exact_id_match:
            # Exact ID match takes priority
            selected_student_obj = exact_id_match
        elif search_students.count() == 1:
            # If search returns exactly one student, show their profile
            selected_student_obj = search_students.first()
        elif search_students.count() > 1:
            # Multiple results - prepare search results info
            search_results_info = {
                'count': search_students.count(),
                'students': search_students[:5]  # Show first 5 matches
            }

    context = {
        'total_logs': total_logs,
        'reviewed_logs': reviewed_logs,
        'pending_logs': pending_logs,
        'approved_logs': approved_logs,
        'rejected_logs': rejected_logs,
        'doctor_names': doctor_names,
        'total_doctors': len(doctor_names),
        'departments': departments,
        'students': students,
        'selected_department': department_filter,
        'selected_student': student_filter,
        'selected_student_obj': selected_student_obj,
        'search_query': search_query,
        'search_results_info': search_results_info,
        'case_types_data': json.dumps(case_types_chart),
        'training_sites_data': json.dumps(training_sites_chart),
        'activity_types_data': json.dumps(activity_types_chart),
        'participation_data': json.dumps(participation_chart),
        'monthly_data': json.dumps(monthly_chart),
        'approval_status_data': json.dumps(approval_status_chart),
    }

    return render(request, 'admin_section/student_report.html', context)


@login_required
def student_report_export(request):
    """Export Student Report as PDF or Excel"""
    if request.user.role != 'admin':
        messages.error(request, "You don't have permission to access this page.")
        return redirect('admin_section:admin_dash')

    export_format = request.GET.get('format', 'pdf')

    # Get filter parameters
    selected_department = request.GET.get('department')
    selected_year = request.GET.get('year')
    selected_group = request.GET.get('group')
    selected_student = request.GET.get('student')

    # Get students with filters
    students = Student.objects.select_related('user', 'group', 'group__log_year').all()

    # Filter by department through logs (since Group doesn't have department)
    if selected_department:
        # Get students who have logs in the selected department
        student_ids_with_dept_logs = StudentLogFormModel.objects.filter(
            department_id=selected_department
        ).values_list('student_id', flat=True).distinct()
        students = students.filter(id__in=student_ids_with_dept_logs)

    if selected_year:
        students = students.filter(group__log_year__year_name=selected_year)
    if selected_group:
        students = students.filter(group_id=selected_group)
    if selected_student:
        students = students.filter(id=selected_student)

    # Prepare data for export
    student_data = []
    for student in students:
        # Get logs for this student
        logs = StudentLogFormModel.objects.filter(student=student)
        total_logs = logs.count()
        reviewed_logs = logs.filter(is_reviewed=True).count()
        pending_logs = total_logs - reviewed_logs

        # Get unique departments
        departments_count = logs.values('department').distinct().count()

        # Get primary department from logs (most frequent department)
        primary_department = 'N/A'
        if logs.exists():
            dept_counts = logs.values('department__name').annotate(count=models.Count('department')).order_by('-count')
            if dept_counts:
                primary_department = dept_counts[0]['department__name']

        student_data.append({
            'name': f"{student.user.first_name} {student.user.last_name}",
            'email': student.user.email,
            'group': student.group.group_name if student.group else 'N/A',
            'department': primary_department,
            'year': student.group.log_year.year_name if student.group and student.group.log_year else 'N/A',
            'total_logs': total_logs,
            'reviewed_logs': reviewed_logs,
            'pending_logs': pending_logs,
            'departments_count': departments_count,
            'review_rate': f"{(reviewed_logs/total_logs*100):.1f}%" if total_logs > 0 else "0%"
        })

    if export_format == 'excel':
        return export_student_excel(student_data, selected_department, selected_year, selected_group, selected_student)
    else:
        return export_student_pdf(student_data, selected_department, selected_year, selected_group, selected_student)


def export_student_excel(student_data, selected_department=None, selected_year=None, selected_group=None, selected_student=None):
    """Export student report to Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.chart import BarChart, Reference
    from django.http import HttpResponse

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Report"

    # Add title
    title = "Student Report"
    if selected_department:
        dept_name = Department.objects.get(id=selected_department).name
        title += f" - {dept_name}"
    if selected_year:
        title += f" - {selected_year}"
    if selected_group:
        group_name = Group.objects.get(id=selected_group).group_name
        title += f" - {group_name}"

    ws['A1'] = title
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:J1')

    # Add headers
    headers = ['Name', 'Email', 'Group', 'Department', 'Year', 'Total Logs', 'Reviewed', 'Pending', 'Departments', 'Review Rate']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Add data
    for row, student in enumerate(student_data, 4):
        ws.cell(row=row, column=1, value=student['name'])
        ws.cell(row=row, column=2, value=student['email'])
        ws.cell(row=row, column=3, value=student['group'])
        ws.cell(row=row, column=4, value=student['department'])
        ws.cell(row=row, column=5, value=student['year'])
        ws.cell(row=row, column=6, value=student['total_logs'])
        ws.cell(row=row, column=7, value=student['reviewed_logs'])
        ws.cell(row=row, column=8, value=student['pending_logs'])
        ws.cell(row=row, column=9, value=student['departments_count'])
        ws.cell(row=row, column=10, value=student['review_rate'])

    # Auto-adjust column widths
    headers = ['Name', 'Email', 'Student ID', 'Group', 'Department', 'Year', 'Total Logs', 'Reviewed Logs', 'Pending Logs', 'Review Rate']
    for col_num in range(1, len(headers) + 1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_num)
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_num)
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="student_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

    wb.save(response)
    return response


def export_student_pdf(student_data, selected_department=None, selected_year=None, selected_group=None, selected_student=None):
    """Export student report to PDF"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from django.http import HttpResponse
    import io

    # Create response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="student_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'

    # Create PDF in landscape mode for better table fit
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []

    # Add header
    add_agu_header(elements)

    # Add title
    title = "Student Report"
    if selected_department:
        dept_name = Department.objects.get(id=selected_department).name
        title += f" - {dept_name}"
    if selected_year:
        title += f" - {selected_year}"
    if selected_group:
        group_name = Group.objects.get(id=selected_group).group_name
        title += f" - {group_name}"

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1,
        textColor=colors.HexColor('#1f2937')
    )

    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 20))

    # Create table data
    table_data = [['Name', 'Email', 'Group', 'Department', 'Year', 'Total Logs', 'Reviewed', 'Pending', 'Review Rate']]
    for student in student_data:
        table_data.append([
            student['name'],
            student['email'],
            student['group'],
            student['department'],
            str(student['year']),
            str(student['total_logs']),
            str(student['reviewed_logs']),
            str(student['pending_logs']),
            student['review_rate']
        ])

    # Create table
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))

    elements.append(table)

    # Add footer
    add_footer_info(elements)

    # Build PDF
    doc.build(elements)

    # Get PDF data
    pdf_data = buffer.getvalue()
    buffer.close()

    response.write(pdf_data)
    return response

@login_required
def tutor_report(request):
    """View for Tutor Report with enhanced dashboard and filtering"""
    if request.user.role != 'admin':
        messages.error(request, "You don't have permission to access this page.")
        return redirect('admin_section:admin_dash')

    # Get filter parameters
    department_filter = request.GET.get('department')
    doctor_filter = request.GET.get('doctor')
    search_query = request.GET.get('q', '').strip()

    # Base queryset for logs supervised by doctors
    logs = StudentLogFormModel.objects.select_related(
        'tutor', 'tutor__user', 'student', 'student__user', 'department',
        'activity_type', 'core_diagnosis', 'training_site'
    ).all()

    # Apply department filter
    if department_filter:
        logs = logs.filter(department_id=department_filter)

    # Apply doctor filter
    if doctor_filter:
        logs = logs.filter(tutor_id=doctor_filter)

    # Apply search filter
    if search_query:
        logs = logs.filter(
            Q(tutor__user__first_name__icontains=search_query) |
            Q(tutor__user__last_name__icontains=search_query) |
            Q(tutor__user__email__icontains=search_query)
        )

    # Calculate summary statistics
    total_logs = logs.count()
    reviewed_logs = logs.filter(is_reviewed=True).count()
    pending_logs = logs.filter(is_reviewed=False).count()
    approved_logs = logs.filter(is_reviewed=True).exclude(reviewer_comments__startswith='REJECTED').count()
    rejected_logs = logs.filter(is_reviewed=True, reviewer_comments__startswith='REJECTED').count()

    # Get unique doctors from filtered logs
    unique_doctor_ids = logs.values_list('tutor', flat=True).distinct()
    doctors = Doctor.objects.filter(id__in=unique_doctor_ids).select_related('user').prefetch_related('departments')
    total_doctors = doctors.count()

    # Case Types Data (using core_diagnosis)
    case_types_data = logs.values('core_diagnosis__name').annotate(
        count=Count('id')
    ).order_by('-count')

    case_types_chart = {
        'labels': [item['core_diagnosis__name'] or 'Unknown' for item in case_types_data],
        'data': [item['count'] for item in case_types_data]
    }

    # Diagnosis Types Data (same as case types but can be filtered)
    diagnosis_types_chart = case_types_chart.copy()

    # Activity Types Data
    activity_types_data = logs.values('activity_type__name').annotate(
        count=Count('id')
    ).order_by('-count')

    activity_types_chart = {
        'labels': [item['activity_type__name'] for item in activity_types_data],
        'data': [item['count'] for item in activity_types_data]
    }

    # Monthly Cases Data
    from django.db.models.functions import TruncMonth
    monthly_data = logs.annotate(
        month=TruncMonth('date')
    ).values('month').annotate(
        count=Count('id')
    ).order_by('month')

    monthly_chart = {
        'labels': [item['month'].strftime('%b %Y') for item in monthly_data],
        'data': [item['count'] for item in monthly_data]
    }

    # Supervision Types Data (based on training sites or supervision levels)
    supervision_types_data = logs.values('training_site__name').annotate(
        count=Count('id')
    ).order_by('-count')

    supervision_types_chart = {
        'labels': [item['training_site__name'] or 'Unknown' for item in supervision_types_data],
        'data': [item['count'] for item in supervision_types_data]
    }

    # Approval Status Data
    approval_status_chart = {
        'approved': approved_logs,
        'pending': pending_logs,
        'rejected': rejected_logs
    }

    # Get filter options
    departments = Department.objects.all().order_by('name')
    all_doctors = Doctor.objects.select_related('user').prefetch_related('departments').all().order_by('user__first_name', 'user__last_name')

    # Handle doctor search and profile display
    selected_doctor_obj = None
    search_results_info = None

    if doctor_filter:
        try:
            selected_doctor_obj = Doctor.objects.select_related('user').prefetch_related('departments').get(id=doctor_filter)
        except Doctor.DoesNotExist:
            pass
    elif search_query:
        # Search for doctors
        search_doctors = Doctor.objects.select_related('user').prefetch_related('departments').filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(user__email__icontains=search_query)
        )

        # Check for exact email match first
        exact_email_match = Doctor.objects.select_related('user').prefetch_related('departments').filter(
            user__email__iexact=search_query
        ).first()

        if exact_email_match:
            selected_doctor_obj = exact_email_match
        elif search_doctors.count() == 1:
            selected_doctor_obj = search_doctors.first()
        elif search_doctors.count() > 1:
            search_results_info = {
                'count': search_doctors.count(),
                'doctors': search_doctors[:5]
            }

    context = {
        'total_logs': total_logs,
        'reviewed_logs': reviewed_logs,
        'pending_logs': pending_logs,
        'approved_logs': approved_logs,
        'rejected_logs': rejected_logs,
        'total_doctors': total_doctors,
        'departments': departments,
        'doctors': all_doctors,
        'selected_department': department_filter,
        'selected_doctor': doctor_filter,
        'selected_doctor_obj': selected_doctor_obj,
        'search_query': search_query,
        'search_results_info': search_results_info,
        'case_types_data': json.dumps(case_types_chart),
        'diagnosis_types_data': json.dumps(diagnosis_types_chart),
        'activity_types_data': json.dumps(activity_types_chart),
        'supervision_data': json.dumps(supervision_types_chart),
        'monthly_data': json.dumps(monthly_chart),
        'approval_status_data': json.dumps(approval_status_chart),
    }

    return render(request, 'admin_section/tutor_report.html', context)


@login_required
def tutor_report_export(request):
    """Export Tutor Report as PDF or Excel"""
    if request.user.role != 'admin':
        messages.error(request, "You don't have permission to access this page.")
        return redirect('admin_section:admin_dash')

    export_format = request.GET.get('format', 'pdf')

    # Get filter parameters
    selected_department = request.GET.get('department')
    selected_year = request.GET.get('year')

    # Get doctors (tutors) with filters
    doctors = Doctor.objects.select_related('user').prefetch_related('departments').all()

    if selected_department:
        doctors = doctors.filter(departments__id=selected_department)

    # Prepare data for export
    tutor_data = []
    for doctor in doctors:
        # Get logs reviewed by this doctor
        reviewed_logs = StudentLogFormModel.objects.filter(reviewed_by=doctor.user)
        if selected_year:
            reviewed_logs = reviewed_logs.filter(student__group__log_year__year_name=selected_year)

        # Get unique students this doctor has reviewed
        unique_students = reviewed_logs.values('student').distinct().count()

        # Get unique departments this doctor has reviewed logs for
        unique_departments = reviewed_logs.values('department').distinct().count()

        # Calculate average review time (if available)
        total_reviews = reviewed_logs.count()

        # Get doctor's departments (many-to-many relationship)
        doctor_departments = doctor.departments.all()
        department_names = ', '.join([dept.name for dept in doctor_departments]) if doctor_departments.exists() else 'N/A'

        tutor_data.append({
            'name': f"{doctor.user.first_name} {doctor.user.last_name}",
            'email': doctor.user.email,
            'department': department_names,
            'total_reviews': total_reviews,
            'unique_students': unique_students,
            'unique_departments': unique_departments,
            'specialization': doctor.user.speciality or 'N/A',
            'phone': doctor.user.phone_no or 'N/A'
        })

    if export_format == 'excel':
        return export_tutor_excel(tutor_data, selected_department, selected_year)
    else:
        return export_tutor_pdf(tutor_data, selected_department, selected_year)


def export_tutor_excel(tutor_data, selected_department=None, selected_year=None):
    """Export tutor report to Excel"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.chart import BarChart, Reference
    from django.http import HttpResponse

    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tutor Report"

    # Add title
    title = "Tutor Report"
    if selected_department:
        dept_name = Department.objects.get(id=selected_department).name
        title += f" - {dept_name}"
    if selected_year:
        title += f" - {selected_year}"

    ws['A1'] = title
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:H1')

    # Add headers
    headers = ['Name', 'Email', 'Department', 'Specialization', 'Phone', 'Total Reviews', 'Students Reviewed', 'Departments']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Add data
    for row, tutor in enumerate(tutor_data, 4):
        ws.cell(row=row, column=1, value=tutor['name'])
        ws.cell(row=row, column=2, value=tutor['email'])
        ws.cell(row=row, column=3, value=tutor['department'])
        ws.cell(row=row, column=4, value=tutor['specialization'])
        ws.cell(row=row, column=5, value=tutor['phone'])
        ws.cell(row=row, column=6, value=tutor['total_reviews'])
        ws.cell(row=row, column=7, value=tutor['unique_students'])
        ws.cell(row=row, column=8, value=tutor['unique_departments'])

    # Add chart
    if len(tutor_data) > 0:
        chart = BarChart()
        chart.title = "Tutor Review Statistics"
        chart.x_axis.title = "Tutors"
        chart.y_axis.title = "Count"

        # Data for chart
        data = Reference(ws, min_col=6, min_row=3, max_col=8, max_row=len(tutor_data) + 3)
        categories = Reference(ws, min_col=1, min_row=4, max_row=len(tutor_data) + 3)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)

        # Add chart to worksheet
        ws.add_chart(chart, "J3")

    # Auto-adjust column widths
    for col_num in range(1, len(headers) + 1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_num)
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_num)
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="tutor_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

    wb.save(response)
    return response


def export_tutor_pdf(tutor_data, selected_department=None, selected_year=None):
    """Export tutor report to PDF"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    from django.http import HttpResponse
    import io

    # Create response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="tutor_report_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'

    # Create PDF in landscape mode for better table fit
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []

    # Add header
    add_agu_header(elements)

    # Add title
    title = "Tutor Report"
    if selected_department:
        dept_name = Department.objects.get(id=selected_department).name
        title += f" - {dept_name}"
    if selected_year:
        title += f" - {selected_year}"

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1,
        textColor=colors.HexColor('#1f2937')
    )

    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 20))

    # Create table data
    table_data = [['Name', 'Email', 'Department', 'Specialization', 'Total Reviews', 'Students Reviewed', 'Departments']]
    for tutor in tutor_data:
        table_data.append([
            tutor['name'],
            tutor['email'],
            tutor['department'],
            tutor['specialization'],
            str(tutor['total_reviews']),
            str(tutor['unique_students']),
            str(tutor['unique_departments'])
        ])

    # Create table
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))

    elements.append(table)

    # Add footer
    add_footer_info(elements)

    # Build PDF
    doc.build(elements)

    # Get PDF data
    pdf_data = buffer.getvalue()
    buffer.close()

    response.write(pdf_data)
    return response

@login_required
def bulk_add_users(request):
    if request.user.role != 'admin':
        messages.error(request, "You don't have permission to access this page.")
        return redirect('login')

    if request.method == 'POST':
        form = BulkUserUploadForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = request.FILES['csv_file']

            # Validate file size (5MB limit)
            if csv_file.size > 5 * 1024 * 1024:
                messages.error(request, 'File size must be less than 5MB')
                return redirect('admin_section:bulk_add_users')

            try:
                decoded_file = csv_file.read().decode('utf-8')
            except UnicodeDecodeError:
                messages.error(request, 'Please upload a valid CSV file')
                return redirect('admin_section:bulk_add_users')

            io_string = io.StringIO(decoded_file)
            reader = csv.DictReader(io_string)

            # Basic required fields for all user types
            required_fields = [
                'username', 'first_name', 'last_name', 'email',
                'password', 'role', 'city', 'country', 'phone_no'
            ]

            # Validate headers
            headers = reader.fieldnames
            if not headers or not all(field in headers for field in required_fields):
                messages.error(request, 'CSV file must contain all required fields')
                return redirect('admin_section:bulk_add_users')

            success_count = 0
            error_count = 0
            error_messages = []

            for row in reader:
                try:
                    with transaction.atomic():
                        # Basic validation
                        username = row.get('username', '').strip()
                        email = row.get('email', '').strip()
                        role = row.get('role', '').strip().lower()

                        if not username or not email or not role:
                            raise ValueError("Username, email and role are required")

                        if CustomUser.objects.filter(username=username).exists():
                            raise ValueError(f"Username '{username}' already exists")

                        if CustomUser.objects.filter(email=email).exists():
                            raise ValueError(f"Email '{email}' already exists")

                        if role not in ['admin', 'student', 'doctor', 'staff']:
                            raise ValueError(f"Invalid role: {role}")

                        # Create user
                        user = CustomUser.objects.create(
                            username=username,
                            email=email,
                            password=make_password(row['password'].strip()),
                            first_name=row.get('first_name', '').strip(),
                            last_name=row.get('last_name', '').strip(),
                            role=role,
                            phone_no=row.get('phone_no', '').strip(),
                            city=row.get('city', '').strip(),
                            country=row.get('country', '').strip(),
                            bio=row.get('bio', '').strip(),
                            speciality=row.get('speciality', '').strip()
                        )

                        # Create role-specific profile
                        if role == 'student':
                            # Additional validation for student-specific fields
                            student_id = row.get('student_id', '').strip()
                            group_id = row.get('group', '').strip()

                            if not student_id:
                                raise ValueError("Student ID is required for student users")

                            if Student.objects.filter(student_id=student_id).exists():
                                raise ValueError(f"Student ID '{student_id}' already exists")

                            # Create student profile
                            student = Student(user=user, student_id=student_id)

                            # Assign group if provided and valid
                            if group_id:
                                try:
                                    group = Group.objects.get(id=group_id)
                                    student.group = group
                                except Group.DoesNotExist:
                                    # Try to match by group name (B1, B2, A1, etc.)
                                    try:
                                        group = Group.objects.filter(group_name=group_id).first()
                                        if group:
                                            student.group = group
                                    except Exception:
                                        pass

                            student.save()

                        elif role == 'doctor':
                            Doctor.objects.create(user=user)

                        elif role == 'staff':
                            Staff.objects.create(user=user)

                        success_count += 1
                except Exception as e:
                    error_count += 1
                    error_messages.append(f"Row {reader.line_num}: {str(e)}")

            if success_count > 0:
                messages.success(request, f"Successfully added {success_count} users.")
            if error_count > 0:
                messages.warning(request, f"Failed to add {error_count} users. See details below.")

            return render(request, 'admin_section/bulk_add_users.html', {
                'form': form,
                'results': {
                    'success_count': success_count,
                    'error_count': error_count,
                    'error_messages': error_messages[:10],
                    'total_errors': len(error_messages)
                }
            })
    else:
        form = BulkUserUploadForm()

    return render(request, 'admin_section/bulk_add_users.html', {'form': form})

# Create your views here.

@login_required
def date_restrictions(request):
    """View for managing date restrictions for students and doctors"""
    # Check if user is admin
    if request.user.role != 'admin':
        messages.error(request, "You don't have permission to access this page.")
        return redirect('admin_section:admin_dash')

    # Get or create settings
    settings, created = DateRestrictionSettings.objects.get_or_create(pk=1)

    if request.method == 'POST':
        # Process form data
        try:
            # Student settings
            settings.past_days_limit = int(request.POST.get('student_past_days_limit', 7))
            settings.allow_future_dates = 'student_allow_future_dates' in request.POST
            settings.future_days_limit = int(request.POST.get('student_future_days_limit', 0))

            # Doctor review period settings
            settings.doctor_review_enabled = 'doctor_review_enabled' in request.POST
            settings.doctor_review_period = int(request.POST.get('doctor_review_period', 30))
            settings.doctor_notification_days = int(request.POST.get('doctor_notification_days', 3))
            
            # Doctor settings - ensure doctors can mark attendance for today
            settings.doctor_past_days_limit = int(request.POST.get('doctor_past_days_limit', 30))
            settings.doctor_allow_future_dates = 'doctor_allow_future_dates' in request.POST
            settings.doctor_future_days_limit = int(request.POST.get('doctor_future_days_limit', 0))

            # Process allowed days for students
            student_days = []
            for day_value, _ in settings.DAYS_OF_WEEK:
                if f'student_days_{day_value}' in request.POST:
                    student_days.append(str(day_value))

            # If no days selected, default to all days
            settings.allowed_days_for_students = ','.join(student_days) if student_days else '0,1,2,3,4,5,6'

            # Process allowed days for doctors
            doctor_days = []
            for day_value, _ in settings.DAYS_OF_WEEK:
                if f'doctor_days_{day_value}' in request.POST:
                    doctor_days.append(str(day_value))

            # If no days selected, default to all days
            settings.allowed_days_for_doctors = ','.join(doctor_days) if doctor_days else '0,1,2,3,4,5,6'
            
            # Active status
            settings.is_active = 'is_active' in request.POST

            # Attendance tracking setting
            settings.attendance_tracking_enabled = 'attendance_tracking_enabled' in request.POST

            # Set updated by
            settings.updated_by = request.user

            # Save settings
            settings.save()

            messages.success(request, "Date restrictions updated successfully.")
        except Exception as e:
            messages.error(request, f"Error updating date restrictions: {str(e)}")

    # Get unread notifications count for the admin
    unread_notifications_count = AdminNotification.objects.filter(recipient=request.user, is_read=False).count()

    context = {
        'settings': settings,
        'unread_notifications_count': unread_notifications_count,
        'admin_unread_notifications_count': unread_notifications_count,
    }

    return render(request, 'admin_section/date_restrictions_simple.html', context)

@login_required
def admin_dash(request):
    # Get filter parameters
    department_id = request.GET.get('department')
    student_search = request.GET.get('student_search', '')

    # Get all departments and years
    departments = Department.objects.all().order_by('name')
    years = LogYear.objects.all().order_by('-year_name')

    # Base queryset for logs
    logs = StudentLogFormModel.objects.select_related('student', 'student__user', 'department', 'activity_type', 'core_diagnosis').all()
    doctors = Doctor.objects.select_related('user').prefetch_related('departments').all()
    students = Student.objects.select_related('user', 'group').all()
    activities = ActivityType.objects.all()

    # Filter logs by department if selected
    if department_id:
        logs = logs.filter(department_id=department_id)
        doctors = doctors.filter(departments__id=department_id)

    # Get current date and start of month
    today = timezone.now()
    start_of_month = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    # Count metrics
    total_logs = logs.count()
    total_doctors = Doctor.objects.count()
    total_student = Student.objects.count()
    total_departments = Department.objects.count()
    total_activities = ActivityType.objects.count()

    # Review status counts
    reviewed_logs = logs.filter(is_reviewed=True)
    pending_logs = logs.filter(is_reviewed=False).count()
    approved_logs = reviewed_logs.exclude(reviewer_comments__startswith='REJECTED').count()
    rejected_logs = reviewed_logs.filter(reviewer_comments__startswith='REJECTED').count()

    # Get recent logs for the table (limited to 10)
    recent_logs = logs.order_by('-created_at')[:10]

    # Department statistics for charts
    department_stats = []
    for dept in departments:
        dept_logs = StudentLogFormModel.objects.filter(department=dept)
        total = dept_logs.count()
        reviewed = dept_logs.filter(is_reviewed=True).count()
        pending = total - reviewed
        department_stats.append({
            'name': dept.name,
            'total': total,
            'reviewed': reviewed,
            'pending': pending,
            'doctors_count': dept.doctors.count()
        })

    # Doctor performance data if department is selected
    doctor_performance = []
    if department_id:
        for doctor in doctors:
            doctor_logs = logs.filter(tutor=doctor)
            reviewed = doctor_logs.filter(is_reviewed=True).count()
            approved = doctor_logs.filter(is_reviewed=True).exclude(reviewer_comments__startswith='REJECTED').count()
            rejected = doctor_logs.filter(is_reviewed=True, reviewer_comments__startswith='REJECTED').count()
            doctor_performance.append({
                'name': doctor.user.get_full_name() or doctor.user.username,
                'reviewed': reviewed,
                'approved': approved,
                'rejected': rejected
            })

    # Student performance search
    student_data = None
    student_performance_data = None
    if student_search:
        # Search for student by name, ID or email
        student_query = students.filter(
            Q(student_id__icontains=student_search) |
            Q(user__email__icontains=student_search) |
            Q(user__first_name__icontains=student_search) |
            Q(user__last_name__icontains=student_search)
        ).first()

        if student_query:
            # Get student logs
            student_logs = logs.filter(student=student_query)
            total_student_logs = student_logs.count()
            approved_student_logs = student_logs.filter(is_reviewed=True).exclude(reviewer_comments__startswith='REJECTED').count()
            pending_student_logs = student_logs.filter(is_reviewed=False).count()
            rejected_student_logs = student_logs.filter(is_reviewed=True, reviewer_comments__startswith='REJECTED').count()

            # Create student data dictionary
            student_data = {
                'name': student_query.user.get_full_name() or student_query.user.username,
                'id': student_query.student_id,
                'email': student_query.user.email,
                'total_logs': total_student_logs,
                'approved_logs': approved_student_logs,
                'pending_logs': pending_student_logs,
                'rejected_logs': rejected_student_logs
            }

            # Get performance by department
            student_departments = student_logs.values('department__name').annotate(
                total=Count('id'),
                approved=Count('id', filter=Q(is_reviewed=True) & ~Q(reviewer_comments__startswith='REJECTED')),
                pending=Count('id', filter=Q(is_reviewed=False)),
                rejected=Count('id', filter=Q(is_reviewed=True) & Q(reviewer_comments__startswith='REJECTED'))
            )

            student_performance_data = []
            for dept in student_departments:
                student_performance_data.append({
                    'department': dept['department__name'],
                    'total': dept['total'],
                    'approved': dept['approved'],
                    'pending': dept['pending'],
                    'rejected': dept['rejected']
                })

    # Prepare chart data
    chart_data = {
        'department_stats': department_stats
    }

    context = {
        'departments': departments,
        'years': years,
        'selected_department': department_id,
        'total_logs': total_logs,
        'total_doctors': total_doctors,
        'total_student': total_student,
        'total_departments': total_departments,
        'total_activities': total_activities,
        'approved_logs': approved_logs,
        'pending_logs': pending_logs,
        'rejected_logs': rejected_logs,
        'recent_logs': recent_logs,
        'chart_data': chart_data,
        'doctor_performance': doctor_performance,
        'student_search': student_search,
        'student_data': student_data,
        'student_performance_data': student_performance_data
    }

    return render(request, "admin_section/admin_dash.html", context)

def calculate_approval_rate(logs):
    reviewed_logs = logs.filter(is_reviewed=True)
    total_reviewed = reviewed_logs.count()
    if total_reviewed == 0:
        return 0
    approved = reviewed_logs.filter(reviewer_comments__startswith='REJECTED').count()
    return round((1 - approved / total_reviewed) * 100)

def get_daily_submissions_data(logs):
    last_7_days = timezone.now() - timedelta(days=7)
    daily_submissions = logs.filter(
        created_at__gte=last_7_days
    ).values('created_at__date').annotate(
        count=Count('id')
    ).order_by('created_at__date')

    return {
        'labels': [d['created_at__date'].strftime('%Y-%m-%d') for d in daily_submissions],
        'data': [d['count'] for d in daily_submissions]
    }

def get_department_stats(departments):
    dept_stats = []
    for dept in departments:
        logs = StudentLogFormModel.objects.filter(department=dept)
        total = logs.count()
        reviewed = logs.filter(is_reviewed=True).count()
        dept_stats.append({
            'name': dept.name,
            'total': total,
            'reviewed': reviewed,
            'pending': total - reviewed,
            'doctors_count': dept.doctors.count()
        })
    return dept_stats

def get_review_status_data(logs):
    total = logs.count()
    reviewed = logs.filter(is_reviewed=True).count()
    return {
        'labels': ['Reviewed', 'Pending'],
        'data': [reviewed, total - reviewed]
    }

def get_monthly_trend_data(logs):
    last_6_months = timezone.now() - timedelta(days=180)
    monthly_data = logs.filter(
        created_at__gte=last_6_months
    ).annotate(
        month=TruncMonth('created_at')
    ).values('month').annotate(
        count=Count('id')
    ).order_by('month')

    return {
        'labels': [d['month'].strftime('%B %Y') for d in monthly_data],
        'data': [d['count'] for d in monthly_data]
    }


@login_required
def get_user_data(request):
    """AJAX endpoint to get user data for the dashboard"""
    user_type = request.GET.get('user_type', '')

    if not user_type or user_type not in ['student', 'doctor', 'staff', 'admin']:
        return JsonResponse({'error': 'Invalid user type'}, status=400)

    data = {'count': 0, 'users': []}

    if user_type == 'student':
        students = Student.objects.select_related('user', 'group').all()[:10]
        data['count'] = Student.objects.count()

        for student in students:
            data['users'].append({
                'id': student.student_id,
                'name': student.user.get_full_name() or student.user.username,
                'email': student.user.email,
                'group': student.group.group_name if student.group else 'No Group'
            })

    elif user_type == 'doctor':
        doctors = Doctor.objects.select_related('user').prefetch_related('departments').all()[:10]
        data['count'] = Doctor.objects.count()

        for doctor in doctors:
            departments = [dept.name for dept in doctor.departments.all()]
            data['users'].append({
                'name': doctor.user.get_full_name() or doctor.user.username,
                'email': doctor.user.email,
                'departments': ', '.join(departments) if departments else 'No Department',
                'speciality': doctor.user.speciality or 'Not specified'
            })

    elif user_type == 'staff':
        staff = Staff.objects.select_related('user').prefetch_related('departments').all()[:10]
        data['count'] = Staff.objects.count()

        for staff_member in staff:
            departments = [dept.name for dept in staff_member.departments.all()]
            data['users'].append({
                'name': staff_member.user.get_full_name() or staff_member.user.username,
                'email': staff_member.user.email,
                'departments': ', '.join(departments) if departments else 'No Department'
            })

    elif user_type == 'admin':
        admins = CustomUser.objects.filter(role='admin')[:10]
        data['count'] = CustomUser.objects.filter(role='admin').count()

        for admin in admins:
            data['users'].append({
                'name': admin.get_full_name() or admin.username,
                'email': admin.email
            })

    return JsonResponse(data)




def download_user_template(request):
    """Download a sample CSV template for user import"""
    response = HttpResponse(content_type='text/csv')
    user_type = request.GET.get('type', 'general')

    if user_type == 'student':
        response['Content-Disposition'] = 'attachment; filename="student_import_template.csv"'

        # Write headers for student template
        headers = [
            'username', 'email', 'password', 'first_name', 'last_name',
            'role', 'student_id', 'group', 'city', 'country', 'phone_no'
        ]
        writer = csv.writer(response)
        writer.writerow(headers)

        # Write sample data for student with different group names
        sample_data = [
            'student1', 'student1@example.com', 'SecurePass123', 'John', 'Student',
            'student', 'STU12345', 'B1', 'New York', 'USA', '1234567890'
        ]
        writer.writerow(sample_data)

        # Add more examples with different groups
        sample_data2 = [
            'student2', 'student2@example.com', 'SecurePass456', 'Jane', 'Student',
            'student', 'STU67890', 'A2', 'London', 'UK', '9876543210'
        ]
        writer.writerow(sample_data2)

        sample_data3 = [
            'student3', 'student3@example.com', 'SecurePass789', 'Alex', 'Student',
            'student', 'STU24680', 'B2', 'Paris', 'France', '5555555555'
        ]
        writer.writerow(sample_data3)

        sample_data4 = [
            'student4', 'student4@example.com', 'SecurePass101', 'Maria', 'Student',
            'student', 'STU13579', 'A1', 'Berlin', 'Germany', '6666666666'
        ]
        writer.writerow(sample_data4)

    elif user_type == 'doctor':
        response['Content-Disposition'] = 'attachment; filename="doctor_import_template.csv"'

        # Write headers for doctor template
        headers = [
            'username', 'email', 'password', 'first_name', 'last_name',
            'role', 'speciality', 'city', 'country', 'phone_no', 'bio'
        ]
        writer = csv.writer(response)
        writer.writerow(headers)

        # Write sample data for doctor
        sample_data = [
            'doctor1', 'doctor1@example.com', 'SecurePass123', 'John', 'Doctor',
            'doctor', 'Cardiology', 'New York', 'USA', '1234567890', 'Experienced cardiologist'
        ]
        writer.writerow(sample_data)

    elif user_type == 'staff':
        response['Content-Disposition'] = 'attachment; filename="staff_import_template.csv"'

        # Write headers for staff template
        headers = [
            'username', 'email', 'password', 'first_name', 'last_name',
            'role', 'city', 'country', 'phone_no'
        ]
        writer = csv.writer(response)
        writer.writerow(headers)

        # Write sample data for staff
        sample_data = [
            'staff1', 'staff1@example.com', 'SecurePass123', 'John', 'Staff',
            'staff', 'New York', 'USA', '1234567890'
        ]
        writer.writerow(sample_data)

    else:
        # Default general template
        response['Content-Disposition'] = 'attachment; filename="user_import_template.csv"'

        # Write headers for general template
        headers = [
            'username', 'email', 'password', 'first_name', 'last_name',
            'role', 'city', 'country', 'phone_no', 'bio', 'speciality'
        ]
        writer = csv.writer(response)
        writer.writerow(headers)

        # Write sample data
        sample_data = [
            'john.doe', 'john@example.com', 'SecurePass123', 'John', 'Doe',
            'doctor', 'New York', 'USA', '1234567890', 'Experienced doctor', 'Cardiology'
        ]
        writer.writerow(sample_data)

    return response

@login_required
def admin_blogs(request):
    """View for listing and managing blog posts"""
    # Check if user is admin
    if request.user.role != 'admin':
        messages.error(request, "You don't have permission to access this page.")
        return redirect('admin_section:admin_dash')

    # Get filter parameters
    category = request.GET.get('category', '')
    search_query = request.GET.get('q', '').strip()

    # Base queryset
    blogs = Blog.objects.all()

    # Apply filters
    if category:
        if category.startswith('new_'):
            category_id = category.replace('new_', '')
            blogs = blogs.filter(category_new_id=category_id)
        else:
            blogs = blogs.filter(category=category, category_new__isnull=True)

    if search_query:
        blogs = blogs.filter(
            models.Q(title__icontains=search_query) |
            models.Q(summary__icontains=search_query) |
            models.Q(content__icontains=search_query)
        )

    # Order by most recent first
    blogs = blogs.order_by('-created_at')

    # Pagination
    paginator = Paginator(blogs, 10)  # 10 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Get all categories (legacy + new)
    all_categories = list(Blog.CATEGORY_CHOICES)
    for cat in BlogCategory.objects.filter(is_active=True):
        all_categories.append((f"new_{cat.id}", cat.name))

    context = {
        'blogs': page_obj,
        'selected_category': category,
        'search_query': search_query,
        'categories': all_categories,
        'blog_categories': BlogCategory.objects.filter(is_active=True),
    }

    return render(request, "admin_section/admin_blogs.html", context)


@login_required
def blog_create(request):
    """View for creating a new blog post"""
    # Check if user is admin
    if request.user.role != 'admin':
        messages.error(request, "You don't have permission to access this page.")
        return redirect('admin_section:admin_dash')

    if request.method == 'POST':
        form = BlogForm(request.POST, request.FILES)
        if form.is_valid():
            blog = form.save(commit=False)
            blog.author = request.user
            blog.save()
            messages.success(request, "Blog post created successfully.")
            return redirect('admin_section:admin_blogs')
    else:
        form = BlogForm()

    context = {
        'form': form,
        'is_create': True,
    }

    return render(request, "admin_section/blog_form.html", context)


@login_required
def blog_edit(request, blog_id):
    """View for editing an existing blog post"""
    # Check if user is admin
    if request.user.role != 'admin':
        messages.error(request, "You don't have permission to access this page.")
        return redirect('admin_section:admin_dash')

    blog = get_object_or_404(Blog, id=blog_id)

    if request.method == 'POST':
        form = BlogForm(request.POST, request.FILES, instance=blog)
        if form.is_valid():
            form.save()
            messages.success(request, "Blog post updated successfully.")
            return redirect('admin_section:admin_blogs')
    else:
        form = BlogForm(instance=blog)

    context = {
        'form': form,
        'blog': blog,
        'is_create': False,
    }

    return render(request, "admin_section/blog_form.html", context)


@login_required
def blog_delete(request, blog_id):
    """View for deleting a blog post"""
    # Check if user is admin
    if request.user.role != 'admin':
        messages.error(request, "You don't have permission to access this page.")
        return redirect('admin_section:admin_dash')

    blog = get_object_or_404(Blog, id=blog_id)

    if request.method == 'POST':
        # Delete associated files
        if blog.featured_image:
            try:
                if os.path.exists(blog.featured_image.path):
                    os.remove(blog.featured_image.path)
            except Exception as e:
                print(f"Error deleting featured image: {e}")

        if blog.attachment:
            try:
                if os.path.exists(blog.attachment.path):
                    os.remove(blog.attachment.path)
            except Exception as e:
                print(f"Error deleting attachment: {e}")

        # Delete the blog post
        blog.delete()
        messages.success(request, "Blog post deleted successfully.")
        return redirect('admin_section:admin_blogs')

    context = {
        'blog': blog,
    }

    return render(request, "admin_section/blog_confirm_delete.html", context)


@login_required
def blog_detail(request, blog_id):
    """View for viewing a blog post details"""
    # Check if user is admin
    if request.user.role != 'admin':
        messages.error(request, "You don't have permission to access this page.")
        return redirect('admin_section:admin_dash')

    blog = get_object_or_404(Blog, id=blog_id)

    context = {
        'blog': blog,
    }

    return render(request, "admin_section/blog_detail.html", context)


@login_required
def blog_categories(request):
    """View for managing blog categories"""
    # Check if user is admin
    if request.user.role != 'admin':
        messages.error(request, "You don't have permission to access this page.")
        return redirect('admin_section:admin_dash')

    if request.method == 'POST':
        form = BlogCategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Blog category created successfully.")
            return redirect('admin_section:blog_categories')
    else:
        form = BlogCategoryForm()

    # Get all categories
    categories = BlogCategory.objects.all().order_by('name')

    # Pagination
    paginator = Paginator(categories, 10)  # 10 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'form': form,
        'categories': page_obj,
    }

    return render(request, "admin_section/blog_categories.html", context)


@login_required
def blog_category_edit(request, category_id):
    """View for editing a blog category"""
    # Check if user is admin
    if request.user.role != 'admin':
        messages.error(request, "You don't have permission to access this page.")
        return redirect('admin_section:admin_dash')

    category = get_object_or_404(BlogCategory, id=category_id)

    if request.method == 'POST':
        form = BlogCategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, "Blog category updated successfully.")
            return redirect('admin_section:blog_categories')
    else:
        form = BlogCategoryForm(instance=category)

    context = {
        'form': form,
        'category': category,
        'is_edit': True,
    }

    return render(request, "admin_section/blog_category_form.html", context)


@login_required
def blog_category_delete(request, category_id):
    """View for deleting a blog category"""
    # Check if user is admin
    if request.user.role != 'admin':
        messages.error(request, "You don't have permission to access this page.")
        return redirect('admin_section:admin_dash')

    category = get_object_or_404(BlogCategory, id=category_id)

    # Check if category is being used by any blogs
    blogs_count = Blog.objects.filter(category_new=category).count()

    if blogs_count > 0:
        messages.error(request, f"Cannot delete category '{category.name}' because it is being used by {blogs_count} blog post(s).")
        return redirect('admin_section:blog_categories')

    if request.method == 'POST':
        category_name = category.name
        category.delete()
        messages.success(request, f"Blog category '{category_name}' deleted successfully.")
        return redirect('admin_section:blog_categories')

    context = {
        'category': category,
        'blogs_count': blogs_count,
    }

    return render(request, "admin_section/blog_category_confirm_delete.html", context)


@login_required
def admin_profile(request):
    # Get the currently logged-in user from the request
    user = request.user

    # Get the profile photo URL
    if user.profile_photo:
        profile_photo = user.profile_photo.url
    else:
        profile_photo = "/media/profiles/default.jpg"  # Default image if no profile photo

    # Get user information
    username = user.username
    first_name = user.first_name
    last_name = user.last_name
    full_name = user.get_full_name() or username
    user_email = user.email
    user_role = user.role
    user_city = user.city
    user_country = user.country
    user_phone = user.phone_no
    user_speciality = user.speciality

    # Prepare the context dictionary with all the necessary data
    data = {
        "username": username,
        "full_name": full_name,
        "profile_photo": profile_photo,
        "user_role": user_role,
        "first_name": first_name,
        "last_name": last_name,
        "user_city": user_city,
        "user_country": user_country,
        "user_phone": user_phone,
        "user_speciality": user_speciality,
        "user_email": user_email,
    }

    return render(request, "admin_section/admin_profile.html", data)


@login_required
def admin_reviews(request):
    # Get filter parameters
    department_id = request.GET.get('department')
    status = request.GET.get('status', 'pending')
    search_query = request.GET.get('q', '').strip()

    # Get all departments for the filter
    all_departments = Department.objects.all()

    # Base queryset - for admin, show logs from all departments
    logs = StudentLogFormModel.objects.all()

    # Filter by review status
    if status == 'pending':
        logs = logs.filter(is_reviewed=False)
    elif status == 'reviewed':
        logs = logs.filter(is_reviewed=True)
    # If 'all' is selected, don't apply any filter

    # Filter by specific department if selected
    if department_id:
        logs = logs.filter(department_id=department_id)

    # Apply search query if provided
    if search_query:
        logs = logs.filter(
            models.Q(student__user__first_name__icontains=search_query) |
            models.Q(student__user__last_name__icontains=search_query) |
            models.Q(student__student_id__icontains=search_query) |
            models.Q(description__icontains=search_query) |
            models.Q(patient_id__icontains=search_query)
        )

    # Order by most recent first
    logs = logs.order_by('-date', '-created_at')

    # Pagination
    paginator = Paginator(logs, 10)  # 10 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Create batch review form
    batch_form = BatchReviewForm()

    context = {
        'logs': page_obj,
        'departments': all_departments,
        'selected_department': department_id,
        'selected_status': status,
        'search_query': search_query,
        'batch_form': batch_form,
    }

    return render(request, "admin_section/admin_reviews.html", context)


@login_required
def admin_support(request):
    # Get ticket type filter
    ticket_type = request.GET.get('type', 'student')
    status_filter = request.GET.get('status', '')

    # Get appropriate tickets based on type
    if ticket_type == 'doctor':
        tickets = DoctorSupportTicket.objects.all().order_by('-date_created')
    else:  # default to student tickets
        tickets = SupportTicket.objects.all().order_by('-date_created')
        ticket_type = 'student'  # ensure valid value

    # Apply status filter if provided
    if status_filter in ['pending', 'solved']:
        tickets = tickets.filter(status=status_filter)

    context = {
        'tickets': tickets,
        'status_filter': status_filter,
        'ticket_type': ticket_type,
    }
    return render(request, "admin_section/admin_support.html", context)


@login_required
def final_records(request):
    return render(request, "admin_section/admin_final_record.html")


@login_required
def update_contact_info(request):
    if request.method == "POST":
        phone = request.POST.get("phone")
        city = request.POST.get("city")
        country = request.POST.get("country")

        try:
            # Update user profile info
            user = request.user
            user.phone_no = phone
            user.city = city
            user.country = country
            user.save()

            # Update session data
            request.session["phone_no"] = phone
            request.session["city"] = city
            request.session["country"] = country
            request.session.modified = True

            return JsonResponse(
                {
                    "success": True,
                    "phone": phone,
                    "city": city,
                    "country": country,
                }
            )
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Invalid request"})


@login_required
def update_profile_photo(request):
    if request.method == "POST" and request.FILES.get("profile_photo"):
        try:
            photo = request.FILES["profile_photo"]

            # Validate file size (120KB = 120 * 1024 bytes)
            max_size = 120 * 1024  # 120KB in bytes
            if photo.size > max_size:
                return JsonResponse({
                    "success": False,
                    "error": f"File size too large. Maximum allowed size is 120KB. Your file is {photo.size // 1024}KB."
                })

            # Validate file type
            allowed_types = ['image/jpeg', 'image/jpg', 'image/png', 'image/gif']
            if photo.content_type not in allowed_types:
                return JsonResponse({
                    "success": False,
                    "error": "Invalid file type. Only JPEG, PNG, and GIF images are allowed."
                })

            user = request.user
            # Delete old profile photo if it exists and it's not the default
            if user.profile_photo and hasattr(user.profile_photo, "path") and "default.jpg" not in user.profile_photo.path:
                try:
                    if os.path.exists(user.profile_photo.path):
                        os.remove(user.profile_photo.path)
                except Exception as e:
                    print(f"Error deleting old profile photo: {e}")

            # Save new profile photo
            user.profile_photo = photo
            user.save()

            return JsonResponse({"success": True, "profile_photo": user.profile_photo.url})
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "No photo provided"})


@login_required
def review_log(request, log_id):
    # Get the log
    log = get_object_or_404(StudentLogFormModel, id=log_id)

    if request.method == 'POST':
        form = LogReviewForm(request.POST, instance=log)
        if form.is_valid():
            # Save the form but don't commit yet
            log_entry = form.save(commit=False)

            # Set review status based on the form choice
            is_approved = form.cleaned_data['is_approved']
            log_entry.is_reviewed = True

            # Add rejection prefix to comments if rejected
            if is_approved == 'False':
                prefix = "REJECTED: "
                if not log_entry.reviewer_comments.startswith(prefix):
                    log_entry.reviewer_comments = prefix + log_entry.reviewer_comments

            # Set review date
            log_entry.review_date = timezone.now()
            log_entry.save()

            messages.success(request, f"Log entry has been {'approved' if is_approved == 'True' else 'rejected'}.")
            return redirect('admin_section:admin_reviews')
    else:
        form = LogReviewForm(instance=log)

    context = {
        'form': form,
        'log': log,
    }

    return render(request, 'admin_section/admin_review_log.html', context)


@login_required
def batch_review(request):
    if request.method != 'POST':
        return redirect('admin_section:admin_reviews')

    form = BatchReviewForm(request.POST)
    if not form.is_valid():
        messages.error(request, "Invalid form submission.")
        return redirect('admin_section:admin_reviews')

    # Get form data
    log_ids = form.cleaned_data['log_ids'].split(',')
    action = form.cleaned_data['action']
    comments = form.cleaned_data['comments']

    # Get logs - admin can review all logs
    logs = StudentLogFormModel.objects.filter(id__in=log_ids)

    # Process each log
    with transaction.atomic():
        for log in logs:
            log.is_reviewed = True
            log.review_date = timezone.now()

            # Set comments based on action
            if action == 'reject':
                prefix = "REJECTED: "
                log.reviewer_comments = prefix + comments if comments else prefix + "Batch rejected"
            else:  # approve
                log.reviewer_comments = comments if comments else "Approved"

            log.save()

    count = logs.count()
    messages.success(request, f"{count} log entries have been {'approved' if action == 'approve' else 'rejected'}.")
    return redirect('admin_section:admin_reviews')


@login_required
def resolve_ticket(request, ticket_id):
    # Check if it's a student or doctor ticket
    ticket_type = request.GET.get('type', 'student')

    if ticket_type == 'doctor':
        ticket = get_object_or_404(DoctorSupportTicket, id=ticket_id)
        form_class = AdminDoctorResponseForm
    else:  # default to student ticket
        ticket = get_object_or_404(SupportTicket, id=ticket_id)
        form_class = AdminResponseForm

    if request.method == 'POST':
        form = form_class(request.POST, instance=ticket)
        if form.is_valid():
            ticket = form.save(commit=False)
            if ticket.status == 'solved':
                ticket.resolved_date = timezone.now()
            ticket.save()
            messages.success(request, f"Ticket '{ticket.subject}' has been updated successfully.")
            return redirect(f"/admin_section/admin_support/?type={ticket_type}")
    else:
        form = form_class(instance=ticket)

    context = {
        'form': form,
        'ticket': ticket,
        'ticket_type': ticket_type,
    }
    return render(request, 'admin_section/resolve_ticket.html', context)


@login_required
def notifications(request):
    # Check if user is admin
    if request.user.role != 'admin':
        messages.error(request, "You don't have permission to access this page.")
        return redirect('login')

    # Get all notifications for this admin
    notifications_list = AdminNotification.objects.filter(recipient=request.user).order_by('-created_at')

    # Apply filters if provided
    filter_param = request.GET.get('filter', '')

    if filter_param == 'unread':
        notifications_list = notifications_list.filter(is_read=False)
    elif filter_param in ['student', 'doctor', 'staff']:
        notifications_list = notifications_list.filter(support_ticket_type=filter_param)
    # 'all' or empty filter shows everything (default behavior)

    # Mark notifications as read if requested
    if request.GET.get('mark_read'):
        notification_id = request.GET.get('mark_read')
        notification = get_object_or_404(AdminNotification, id=notification_id, recipient=request.user)
        notification.is_read = True
        notification.save()

        # Preserve filter when redirecting
        redirect_url = 'admin_section:notifications'
        if filter_param:
            return redirect(f'{redirect_url}?filter={filter_param}')
        return redirect(redirect_url)

    # Mark all as read if requested
    if request.GET.get('mark_all_read'):
        # Apply the same filters to mark only filtered notifications as read
        to_mark = notifications_list.filter(is_read=False)
        count = to_mark.count()
        to_mark.update(is_read=True)

        if count > 0:
            messages.success(request, f"{count} notifications marked as read.")
        else:
            messages.info(request, "No unread notifications to mark as read.")

        # Preserve filter when redirecting
        redirect_url = 'admin_section:notifications'
        if filter_param:
            return redirect(f'{redirect_url}?filter={filter_param}')
        return redirect(redirect_url)

    # View ticket if requested
    if request.GET.get('view_ticket'):
        notification_id = request.GET.get('view_ticket')
        notification = get_object_or_404(AdminNotification, id=notification_id, recipient=request.user)

        # Mark as read
        if not notification.is_read:
            notification.is_read = True
            notification.save()

        # Redirect to the appropriate ticket page
        if notification.ticket_id:
            ticket_type = notification.support_ticket_type
            return redirect(f'/admin_section/resolve_ticket/{notification.ticket_id}/?type={ticket_type}')

    # Get unread count (for display in the UI)
    unread_count = AdminNotification.objects.filter(recipient=request.user, is_read=False).count()

    # Pagination
    paginator = Paginator(notifications_list, 10)  # 10 items per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'notifications': page_obj,
        'unread_count': unread_count,
        'filter': filter_param,
        'total_count': AdminNotification.objects.filter(recipient=request.user).count(),
    }

    return render(request, 'admin_section/notifications.html', context)


@require_POST
@login_required
def delete_all_notifications(request):
    """Handle deletion of all notifications for the current admin"""
    # Check if user is admin
    if request.user.role != 'admin':
        return JsonResponse({
            'success': False,
            'message': 'You do not have permission to perform this action.'
        }, status=403)

    try:
        # Get all notifications for this admin
        notifications = AdminNotification.objects.filter(recipient=request.user)

        if not notifications.exists():
            return JsonResponse({
                'success': False,
                'message': 'No notifications found to delete.'
            })

        # Count notifications before deletion
        notification_count = notifications.count()

        # Perform bulk deletion
        with transaction.atomic():
            notifications.delete()

        return JsonResponse({
            'success': True,
            'message': f'Successfully deleted {notification_count} notification(s).',
            'deleted_count': notification_count
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error occurred during deletion: {str(e)}'
        }, status=500)


@login_required
def bulk_import_users(request):
    if request.user.role != 'admin':
        messages.error(request, "You don't have permission to access this page.")
        return redirect('login')

    if request.method == 'POST':
        form = CSVUploadForm(request.POST, request.FILES)
        if form.is_valid():
            csv_file = request.FILES['csv_file']
            user_type = form.cleaned_data['user_type']

            # Validate file size (5MB limit)
            if csv_file.size > 5 * 1024 * 1024:
                messages.error(request, 'File size must be less than 5MB')
                return redirect('admin_section:bulk_import_users')

            try:
                decoded_file = csv_file.read().decode('utf-8')
            except UnicodeDecodeError:
                messages.error(request, 'Please upload a valid CSV file')
                return redirect('admin_section:bulk_import_users')

            io_string = io.StringIO(decoded_file)
            reader = csv.DictReader(io_string)

            # Basic required fields for all user types
            required_fields = [
                'username', 'email', 'password', 'first_name', 'last_name',
                'city', 'country', 'phone_no'
            ]

            # Add role-specific required fields
            if user_type == 'student':
                required_fields.append('student_id')

            # Validate headers
            headers = reader.fieldnames
            if not headers or not all(field in headers for field in required_fields):
                missing_fields = [field for field in required_fields if field not in headers]
                messages.error(request, f'CSV file is missing required fields: {", ".join(missing_fields)}')
                return redirect('admin_section:bulk_import_users')

            success_count = 0
            error_count = 0
            error_messages = []

            for row in reader:
                try:
                    with transaction.atomic():
                        # Basic validation
                        username = row.get('username', '').strip()
                        email = row.get('email', '').strip()

                        if not username or not email:
                            raise ValueError("Username and email are required")

                        if CustomUser.objects.filter(username=username).exists():
                            raise ValueError(f"Username '{username}' already exists")

                        if CustomUser.objects.filter(email=email).exists():
                            raise ValueError(f"Email '{email}' already exists")

                        # Create user with role based on form selection
                        user = CustomUser.objects.create(
                            username=username,
                            email=email,
                            password=make_password(row['password'].strip()),
                            first_name=row.get('first_name', '').strip(),
                            last_name=row.get('last_name', '').strip(),
                            role=user_type,  # Use the selected role from the form
                            phone_no=row.get('phone_no', '').strip(),
                            city=row.get('city', '').strip(),
                            country=row.get('country', '').strip(),
                            bio=row.get('bio', '').strip(),
                            speciality=row.get('speciality', '').strip()
                        )

                        # Create role-specific profile
                        if user_type == 'student':
                            # Additional validation for student-specific fields
                            student_id = row.get('student_id', '').strip()
                            group_id = row.get('group', '').strip()

                            if not student_id:
                                raise ValueError("Student ID is required for student users")

                            if Student.objects.filter(student_id=student_id).exists():
                                raise ValueError(f"Student ID '{student_id}' already exists")

                            # Create student profile
                            student = Student(user=user, student_id=student_id)

                            # Assign group if provided and valid
                            if group_id:
                                # First try to match by group name (B1, B2, A1, etc.)
                                group = Group.objects.filter(group_name=group_id).first()

                                # If not found by name, try by ID (if it's a number)
                                if not group and group_id.isdigit():
                                    try:
                                        group = Group.objects.get(id=int(group_id))
                                    except Group.DoesNotExist:
                                        pass

                                # If group was found, assign it
                                if group:
                                    student.group = group
                                else:
                                    # Log a warning but don't fail the import
                                    print(f"Warning: Group '{group_id}' not found for student {student_id}")

                            student.save()

                        elif user_type == 'doctor':
                            Doctor.objects.create(user=user)

                        elif user_type == 'staff':
                            Staff.objects.create(user=user)

                        success_count += 1
                except Exception as e:
                    error_count += 1
                    error_messages.append(f"Row {reader.line_num}: {str(e)}")

            if success_count > 0:
                messages.success(request, f"Successfully added {success_count} {user_type}s.")
            if error_count > 0:
                messages.warning(request, f"Failed to add {error_count} {user_type}s. See details below.")

            return render(request, 'admin_section/bulk_import_users.html', {
                'form': form,
                'results': {
                    'success_count': success_count,
                    'error_count': error_count,
                    'error_messages': error_messages[:10],
                    'total_errors': len(error_messages),
                    'user_type': user_type
                }
            })
    else:
        form = CSVUploadForm()

    return render(request, 'admin_section/bulk_import_users.html', {'form': form})

@login_required
def download_sample_csv(request):
    """Download a sample CSV template for student import"""
    # This function is kept for backward compatibility
    # Redirect to the download_user_template function with type=student
    return download_user_template(request)


def send_admin_emails(admin_emails, subject, message):
    """
    Send email notifications to admin users

    Args:
        admin_emails (list): List of admin email addresses
        subject (str): Email subject
        message (str): Email message body
    """
    try:
        from django.core.mail import send_mail
        from django.conf import settings

        # Send email to all admin users
        send_mail(
            subject=subject,
            message=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=admin_emails,
            fail_silently=True,
        )
    except Exception as e:
        # Log the error but don't raise it to prevent disrupting the user experience
        print(f"Error sending admin emails: {str(e)}")
        # In a production environment, you might want to log this to a file or monitoring service


@login_required
def export_users(request):
    """Export users in CSV format based on user type"""
    if request.user.role != 'admin':
        messages.error(request, "You don't have permission to access this page.")
        return redirect('login')

    user_type = request.GET.get('type', 'all')
    today = datetime.now().strftime('%Y-%m-%d')

    # Create HTTP response with CSV content type
    response = HttpResponse(content_type='text/csv')

    # Set filename based on user type
    if user_type == 'student':
        filename = f'student_users_export_{today}.csv'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Get all student users
        users = CustomUser.objects.filter(role='student').select_related('student').order_by('username')

        # Write CSV headers
        writer = csv.writer(response)
        writer.writerow([
            'Username', 'Email', 'First Name', 'Last Name', 'Student ID', 'Group',
            'Phone Number', 'City', 'Country', 'Date Joined', 'Last Login', 'Is Active'
        ])

        # Write data rows
        for user in users:
            student_profile = getattr(user, 'student', None)
            writer.writerow([
                user.username,
                user.email,
                user.first_name,
                user.last_name,
                student_profile.student_id if student_profile else 'N/A',
                student_profile.group.group_name if student_profile and student_profile.group else 'N/A',
                user.phone_no or 'N/A',
                user.city or 'N/A',
                user.country or 'N/A',
                user.date_joined.strftime('%Y-%m-%d %H:%M:%S') if user.date_joined else 'N/A',
                user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else 'Never',
                'Yes' if user.is_active else 'No'
            ])

    elif user_type == 'doctor':
        filename = f'doctor_users_export_{today}.csv'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Get all doctor users
        users = CustomUser.objects.filter(role='doctor').select_related('doctor_profile').prefetch_related('doctor_profile__departments').order_by('username')

        # Write CSV headers
        writer = csv.writer(response)
        writer.writerow([
            'Username', 'Email', 'First Name', 'Last Name', 'Speciality', 'Departments',
            'Phone Number', 'City', 'Country', 'Date Joined', 'Last Login', 'Is Active'
        ])

        # Write data rows
        for user in users:
            doctor_profile = getattr(user, 'doctor_profile', None)
            departments = ', '.join([dept.name for dept in doctor_profile.departments.all()]) if doctor_profile else 'N/A'

            writer.writerow([
                user.username,
                user.email,
                user.first_name,
                user.last_name,
                user.speciality or 'N/A',  # speciality is on CustomUser, not Doctor
                departments,
                user.phone_no or 'N/A',
                user.city or 'N/A',
                user.country or 'N/A',
                user.date_joined.strftime('%Y-%m-%d %H:%M:%S') if user.date_joined else 'N/A',
                user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else 'Never',
                'Yes' if user.is_active else 'No'
            ])

    elif user_type == 'staff':
        filename = f'staff_users_export_{today}.csv'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Get all staff users
        users = CustomUser.objects.filter(role='staff').select_related('staff_profile').order_by('username')

        # Write CSV headers
        writer = csv.writer(response)
        writer.writerow([
            'Username', 'Email', 'First Name', 'Last Name', 'Phone Number', 'City', 'Country',
            'Date Joined', 'Last Login', 'Is Active'
        ])

        # Write data rows
        for user in users:
            writer.writerow([
                user.username,
                user.email,
                user.first_name,
                user.last_name,
                user.phone_no or 'N/A',
                user.city or 'N/A',
                user.country or 'N/A',
                user.date_joined.strftime('%Y-%m-%d %H:%M:%S') if user.date_joined else 'N/A',
                user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else 'Never',
                'Yes' if user.is_active else 'No'
            ])

    else:  # all users
        filename = f'all_users_export_{today}.csv'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Get all users
        users = CustomUser.objects.all().select_related('student', 'doctor_profile', 'staff_profile').prefetch_related('doctor_profile__departments').order_by('role', 'username')

        # Write CSV headers
        writer = csv.writer(response)
        writer.writerow([
            'Username', 'Email', 'First Name', 'Last Name', 'Role', 'Student ID', 'Speciality', 'Departments',
            'Group', 'Phone Number', 'City', 'Country', 'Date Joined', 'Last Login', 'Is Active'
        ])

        # Write data rows
        for user in users:
            student_profile = getattr(user, 'student', None)
            doctor_profile = getattr(user, 'doctor_profile', None)
            departments = ', '.join([dept.name for dept in doctor_profile.departments.all()]) if doctor_profile else 'N/A'

            writer.writerow([
                user.username,
                user.email,
                user.first_name,
                user.last_name,
                user.role.title(),
                student_profile.student_id if student_profile else 'N/A',
                user.speciality or 'N/A',  # speciality is on CustomUser, not Doctor
                departments,
                student_profile.group.group_name if student_profile and student_profile.group else 'N/A',
                user.phone_no or 'N/A',
                user.city or 'N/A',
                user.country or 'N/A',
                user.date_joined.strftime('%Y-%m-%d %H:%M:%S') if user.date_joined else 'N/A',
                user.last_login.strftime('%Y-%m-%d %H:%M:%S') if user.last_login else 'Never',
                'Yes' if user.is_active else 'No'
            ])

    return response


@login_required
def export_department_logs(request):
    """Export all department logs filtered by year and department in PDF or Excel format"""
    # Check if user is admin
    if request.user.role != 'admin':
        messages.error(request, "You don't have permission to access this page.")
        return redirect('admin_section:admin_dash')

    # Get filter parameters
    export_format = request.GET.get('format', 'pdf').lower()
    year_ids = request.GET.getlist('years')  # Get multiple years
    department_id = request.GET.get('department')

    # Base queryset for logs
    logs = StudentLogFormModel.objects.select_related(
        'student', 'student__user', 'student__group', 'student__group__log_year',
        'department', 'activity_type', 'core_diagnosis', 'tutor', 'tutor__user'
    ).all()

    # Apply filters
    if year_ids:
        logs = logs.filter(student__group__log_year_id__in=year_ids)

    if department_id:
        logs = logs.filter(department_id=department_id)

    # Order by most recent first
    logs = logs.order_by('-date', '-created_at')

    # Prepare filename with timestamp
    timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
    filename_base = f"department_logs_{timestamp}"

    if export_format == 'pdf':
        return export_department_logs_pdf(logs, filename_base, request.user, year_ids, department_id)
    elif export_format == 'excel':
        return export_department_logs_excel(logs, filename_base, year_ids, department_id)
    elif export_format == 'csv':
        return export_department_logs_csv(logs, filename_base, year_ids, department_id)
    else:
        # Default to PDF if format is not recognized
        return export_department_logs_pdf(logs, filename_base, request.user, year_ids, department_id)


def export_department_logs_pdf(logs, filename_base, admin_user, year_ids=None, department_id=None):
    """Export department logs as PDF file with AGU logo"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'

    # Create a buffer for the PDF
    buffer = io.BytesIO()

    # Create the PDF document with landscape orientation for better table fit
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    elements = []

    # Add AGU header with logo and university name
    elements = add_agu_header(elements, "Department Logs Report")

    # Get custom styles
    custom_styles = get_common_styles()

    # Add filter information
    admin_name = admin_user.get_full_name() or admin_user.username
    elements.append(Paragraph(f"Generated by: {admin_name}", custom_styles['subtitle']))

    # Add filter details
    filter_info = []
    if year_ids:
        try:
            years = LogYear.objects.filter(id__in=year_ids).values_list('year_name', flat=True)
            if years:
                year_names = list(years)
                if len(year_names) == 1:
                    filter_info.append(f"Academic Year: {year_names[0]}")
                else:
                    filter_info.append(f"Academic Years: {', '.join(year_names)}")
        except Exception:
            pass

    if department_id:
        try:
            department = Department.objects.get(id=department_id)
            filter_info.append(f"Department: {department.name}")
        except Department.DoesNotExist:
            pass

    if not filter_info:
        filter_info.append("All Years and Departments")

    elements.append(Paragraph(f"Filters: {', '.join(filter_info)}", custom_styles['normal']))
    elements.append(Paragraph(f"Total Records: {logs.count()}", custom_styles['normal']))
    elements.append(Spacer(1, 0.3*inch))

    if logs.exists():
        # Create table data
        table_data = [
            ['Student ID', 'Student Name', 'Department', 'Activity', 'Date', 'Status', 'Tutor', 'Review Date']
        ]

        for log in logs:
            # Determine status
            if log.is_reviewed:
                if log.reviewer_comments and log.reviewer_comments.startswith('REJECTED'):
                    status = 'Rejected'
                else:
                    status = 'Approved'
            else:
                status = 'Pending'

            table_data.append([
                log.student.student_id,
                log.student.user.get_full_name() or log.student.user.username,
                log.department.name,
                log.activity_type.name if log.activity_type else 'N/A',
                log.date.strftime('%Y-%m-%d'),
                status,
                log.tutor.user.get_full_name() if log.tutor else 'N/A',
                log.review_date.strftime('%Y-%m-%d') if log.review_date else 'N/A'
            ])

        # Create table
        table = Table(table_data)

        # Define table style
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ])

        table.setStyle(table_style)
        elements.append(table)
    else:
        elements.append(Paragraph("No logs found for the selected criteria.", custom_styles['normal']))

    # Add footer information
    elements = add_footer_info(
        elements,
        generated_by=admin_name,
        export_date=timezone.now().strftime('%Y-%m-%d %H:%M:%S')
    )

    # Build the PDF
    doc.build(elements)

    # Get the value of the buffer and write it to the response
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)

    return response


def export_department_logs_excel(logs, filename_base, year_ids=None, department_id=None):
    """Export department logs as Excel file with AGU logo using openpyxl directly"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    # Create a new workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Logs"

    # Add summary information first
    total_logs = logs.count()
    approved_logs = logs.filter(is_reviewed=True).exclude(reviewer_comments__startswith='REJECTED').count()
    pending_logs = logs.filter(is_reviewed=False).count()
    rejected_logs = logs.filter(is_reviewed=True, reviewer_comments__startswith='REJECTED').count()

    # Define styles
    title_font = Font(bold=True, size=16)
    header_font = Font(bold=True, size=12)
    bold_font = Font(bold=True)

    # Add AGU header
    ws['A1'] = 'Arabian Gulf University - Student Logs Export'
    ws['A1'].font = title_font
    ws.merge_cells('A1:M1')

    # Add export information
    row = 3
    ws[f'A{row}'] = 'Export Date:'
    ws[f'B{row}'] = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
    ws[f'A{row}'].font = bold_font

    row += 1
    # Add filter information
    if year_ids:
        try:
            years = LogYear.objects.filter(id__in=year_ids).values_list('year_name', flat=True)
            if years:
                year_names = list(years)
                if len(year_names) == 1:
                    ws[f'A{row}'] = 'Academic Year:'
                    ws[f'B{row}'] = year_names[0]
                else:
                    ws[f'A{row}'] = 'Academic Years:'
                    ws[f'B{row}'] = ', '.join(year_names)
        except Exception:
            pass
    else:
        ws[f'A{row}'] = 'Academic Years:'
        ws[f'B{row}'] = 'All Years'
    ws[f'A{row}'].font = bold_font

    row += 1
    if department_id:
        try:
            department = Department.objects.get(id=department_id)
            ws[f'A{row}'] = 'Department:'
            ws[f'B{row}'] = department.name
        except Department.DoesNotExist:
            ws[f'A{row}'] = 'Department:'
            ws[f'B{row}'] = 'All Departments'
    else:
        ws[f'A{row}'] = 'Department:'
        ws[f'B{row}'] = 'All Departments'
    ws[f'A{row}'].font = bold_font

    # Add summary statistics
    row += 2
    ws[f'A{row}'] = 'Summary Statistics:'
    ws[f'A{row}'].font = header_font

    row += 1
    ws[f'A{row}'] = 'Total Records:'
    ws[f'B{row}'] = total_logs
    ws[f'A{row}'].font = bold_font

    row += 1
    ws[f'A{row}'] = 'Approved Logs:'
    ws[f'B{row}'] = approved_logs
    ws[f'A{row}'].font = bold_font

    row += 1
    ws[f'A{row}'] = 'Pending Logs:'
    ws[f'B{row}'] = pending_logs
    ws[f'A{row}'].font = bold_font

    row += 1
    ws[f'A{row}'] = 'Rejected Logs:'
    ws[f'B{row}'] = rejected_logs
    ws[f'A{row}'].font = bold_font

    # Add data headers
    row += 3
    headers = [
        'Student ID', 'Student Name', 'Email', 'Group', 'Department',
        'Activity Type', 'Core Diagnosis', 'Date', 'Status', 'Tutor',
        'Review Date', 'Reviewer Comments', 'Created At'
    ]

    # Style the header row
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font_white = Font(bold=True, color='FFFFFF')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Add data rows
    for log in logs:
        row += 1

        # Determine status
        if log.is_reviewed:
            if log.reviewer_comments and log.reviewer_comments.startswith('REJECTED'):
                status = 'Rejected'
            else:
                status = 'Approved'
        else:
            status = 'Pending'

        data_row = [
            str(log.student.student_id) if log.student.student_id else '',
            log.student.user.get_full_name() or log.student.user.username or '',
            log.student.user.email or '',
            log.student.group.group_name if log.student.group else 'N/A',
            log.department.name if log.department else 'N/A',
            log.activity_type.name if log.activity_type else 'N/A',
            log.core_diagnosis.name if log.core_diagnosis else 'N/A',
            log.date.strftime('%Y-%m-%d') if log.date else '',
            status,
            log.tutor.user.get_full_name() if log.tutor else 'N/A',
            log.review_date.strftime('%Y-%m-%d') if log.review_date else 'N/A',
            (log.reviewer_comments[:100] + '...' if log.reviewer_comments and len(log.reviewer_comments) > 100 else (log.reviewer_comments or '')),
            log.created_at.strftime('%Y-%m-%d %H:%M:%S') if log.created_at else ''
        ]

        for col, value in enumerate(data_row, 1):
            ws.cell(row=row, column=col, value=value)

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row_cells in ws[column_letter]:
            try:
                if len(str(row_cells.value)) > max_length:
                    max_length = len(str(row_cells.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Create HTTP response
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'

    return response


def export_department_logs_csv(logs, filename_base, year_ids=None, department_id=None):
    """Export department logs as CSV file"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.csv"'

    writer = csv.writer(response)

    # Add AGU header information
    writer.writerow(['Arabian Gulf University - Department Logs Export'])
    writer.writerow([''])
    writer.writerow(['Export Date:', timezone.now().strftime('%Y-%m-%d %H:%M:%S')])

    # Add filter information
    if year_ids:
        try:
            years = LogYear.objects.filter(id__in=year_ids).values_list('year_name', flat=True)
            if years:
                year_names = list(years)
                if len(year_names) == 1:
                    writer.writerow(['Academic Year:', year_names[0]])
                else:
                    writer.writerow(['Academic Years:', ', '.join(year_names)])
        except Exception:
            pass
    else:
        writer.writerow(['Academic Years:', 'All Years'])

    if department_id:
        try:
            department = Department.objects.get(id=department_id)
            writer.writerow(['Department:', department.name])
        except Department.DoesNotExist:
            writer.writerow(['Department:', 'All Departments'])
    else:
        writer.writerow(['Department:', 'All Departments'])

    # Add summary statistics
    total_logs = logs.count()
    approved_logs = logs.filter(is_reviewed=True).exclude(reviewer_comments__startswith='REJECTED').count()
    pending_logs = logs.filter(is_reviewed=False).count()
    rejected_logs = logs.filter(is_reviewed=True, reviewer_comments__startswith='REJECTED').count()

    writer.writerow([''])
    writer.writerow(['Summary Statistics:'])
    writer.writerow(['Total Records:', total_logs])
    writer.writerow(['Approved Logs:', approved_logs])
    writer.writerow(['Pending Logs:', pending_logs])
    writer.writerow(['Rejected Logs:', rejected_logs])
    writer.writerow([''])
    writer.writerow([''])

    # Write header row for data
    writer.writerow([
        'Student ID', 'Student Name', 'Email', 'Group', 'Department',
        'Activity Type', 'Core Diagnosis', 'Date', 'Status', 'Tutor',
        'Review Date', 'Reviewer Comments', 'Created At'
    ])

    # Write data rows
    for log in logs:
        # Determine status
        if log.is_reviewed:
            if log.reviewer_comments and log.reviewer_comments.startswith('REJECTED'):
                status = 'Rejected'
            else:
                status = 'Approved'
        else:
            status = 'Pending'

        writer.writerow([
            log.student.student_id if log.student.student_id else '',
            log.student.user.get_full_name() or log.student.user.username or '',
            log.student.user.email or '',
            log.student.group.group_name if log.student.group else 'N/A',
            log.department.name if log.department else 'N/A',
            log.activity_type.name if log.activity_type else 'N/A',
            log.core_diagnosis.name if log.core_diagnosis else 'N/A',
            log.date.strftime('%Y-%m-%d') if log.date else '',
            status,
            log.tutor.user.get_full_name() if log.tutor else 'N/A',
            log.review_date.strftime('%Y-%m-%d') if log.review_date else 'N/A',
            log.reviewer_comments or '',
            log.created_at.strftime('%Y-%m-%d %H:%M:%S') if log.created_at else ''
        ])

    return response
